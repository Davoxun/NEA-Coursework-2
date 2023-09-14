from openpyxl import load_workbook, Workbook
import sys
import re
import smtplib
import random
import datetime
from email.message import EmailMessage
from string import Template
from openpyxl.utils import get_column_letter
import sqlite3 #these are all of the libraries which are being used in this program#

with sqlite3.connect("Appointment Details.db") as db:#this connects to the appointments booked database#
  cursor = db.cursor()#allows for queries to be performed on the database#
  
cursor.execute("PRAGMA foreign_keys = 1")
cursor.execute("""CREATE TABLE IF NOT EXISTS Clients(id integer PRIMARY KEY NOT NULL,
name text NOT NULL,              
postcode text NOT NULL);
""") #creates Clients table with 3 fields which are to to do with the client's personal details#

cursor.execute("""CREATE TABLE IF NOT EXISTS GPs (time text PRIMARY KEY NOT NULL,          
gp_pcode text NOT NULL,
date text NOT NULL);""")#creates GPs table which contains 3 fields; this is everything to do with the GP#

cursor.execute("""CREATE TABLE IF NOT EXISTS Appointments (reference integer PRIMARY KEY NOT NULL,
id integer NOT NULL,
time text NOT NULL,
FOREIGN KEY (id) REFERENCES Clients (id)
FOREIGN KEY (time) REFERENCES GPs(time));""")#creates Appointments table with 4 fields (there are 2 foreign keys which are the primary keys of the other tables)#


wb = load_workbook("GP Database.xlsx")#loads excel workbook with GP details(such as name, postcode, etc)#


class Appointment:#relates to everything concerning the client's appointment#
    def __init__(self):#these are the attributes being initialised to the class#
        self.name = "EMPTY"
        self.postcode = "EMPTY"
        self.__ref_num = "EMPTY"
        self.gp_pc = "EMPTY"
        self.time = "EMPTY"
        self.day = "EMPTY"
        self.ID = "EMPTY"
        self.email = "EMPTY"

    def get_ref_num(): #method which returns the user's reference code#
      return Appointment.__ref_num
      
    def unique_id():#method which assigns a unique id to each user#
      Appointment.ID = random.randint(1,1000000000)
      return Appointment.ID
      
    def book_appointment():#books appointment for client#
        print()
        Appointment.ID = int(Appointment.unique_id())
        try:
          confirm = input("Are you sure that you would like to continue with this option(Y/N): ").capitalize()#asks user if they are sure with their option#
        except:
          confirm = input("Invalid entry, please try again:")#asks user again if an invalid entry is made#
        if confirm == "N":
            main_menu()#if they made a mistake, the menu will be displayed for them to choose again#

        full_name = input("Enter your full name: ")#asks user for their full name#
        Appointment.name = full_name#assigns appointment name attribute to user's input#
        postcode = input("Enter your postcode: ").upper()#asks user for their postcode#
        Appointment.postcode = postcode#assigns appointment postcode attribute to user's input#
        cursor.execute("""INSERT INTO Clients(id,name,postcode) VALUES(?,?,?)""", [Appointment.ID,full_name, postcode])#inserts id, name, postcode to Clients table#
        db.commit()      
        f2 = str(postcode[0])#assigns first character of user's postcode to variable#
        ws = wb.active#allows GP database to be accessed#
        ans = {}#empty dictionary for GP postcodes to be appended to#
        count = 1
        for row in range(1,25):#out of the 24 GPs in the database#
            for col in range(3,4):#focuses on the 3rd row which is where all the postcodes are located#
                char = get_column_letter(col)#uses get column letter method to get the 3rd row of the database#
                pc = ws[char + str(row)].value#all the GP postcodes are assigned to variable pc#
                match = re.search(r"\b"+f2,pc)#a search is run on variable pc to find any postcodes which have the same first postcode letter as each other#
                if match is not None:#if there is a match#
                    ans[count] = pc#then that postcode is added to the empty dictionary#
                    count = count + 1#count is incremented by 1#
        print(*[str(n)+ ":" + str(p) for n,p in ans.items()])#prints the dictionary after the loop is finished#

        
        try:
          pick_pc = int(input("Please pick a location: "))#asks user to pick GP postcode#
        except:
          pick_pc = int(input("Invalid entry, please try again:"))#asks the user to enter again if an invalid input is made#
        user_gp = ans[pick_pc]#assigns user's gp postcode to user's option#
        Appointment.gp_pc = user_gp#assigns user's gp postcode to gp_pc attribute#

        ClientTimetable = Timetable()#instantiates Timetable class#
        AppointmentDate = ClientTimetable.Date()#Invokes Date method on Timetable class to retrieve client's appointment date#
        Appointment.day = AppointmentDate#assigns client's date to day attribute#
        AppointmentTime = ClientTimetable.Time()#Invokes Time method on Timetable class to retrieve client's appointment time#
        Appointment.time = AppointmentTime#assigns client's time to time attribute#
        
        email_address = input("Enter your email address: ")#asks user to enter email address#
        email_format = re.compile(r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)")#compiles a regular expression pattern for email addresses#
        validation = email_format.search(email_address)#searches user's email to see if it matches the format of an email address#
        print(validation)
        while validation is None:#if there is no match#
            email_address = input("Invalid email address, please try again: ")#then it asks user to enter email address again#
            validation = email_format.search(email_address)
            print(validation)#then checks to validate the user's input#
            
        email = EmailMessage()#invokes the email message class#
        email['from'] = 'NEA Vaccine Appointment'#this is where the email address is being sent from
        email['to'] = email_address#where email is being sent to#
        email['subject'] = 'Confirmation Email'#subject of the email#
        
        ref_code = str(random.randint(1,1000000))#generates a random number as user's reference
        Appointment.__ref_num = ref_code#assigns refernce code to ref_num attribute#
      
        cursor.execute("""INSERT INTO GPs(time,gp_pcode,date) VALUES(?,?,?)""",[AppointmentTime, user_gp, AppointmentDate])
        db.commit()#inserts the GP postcode, appointment time and appointment date into the GPs table#
      
        cursor.execute("""INSERT INTO Appointments(id,reference,time) VALUES(?,?,?)""",[Appointment.ID,ref_code,AppointmentTime])
        db.commit()#inserts the user's id, GP postcode and reference number into Appointments table#

        string = Template('Thank you for booking an appointment! Your reference code is $ref_code')#body of the email being sent#
        email.set_content(string.substitute(dict(ref_code= ref_code)))

        with smtplib.SMTP(host='smtp.gmail.com', port=587) as smtp:#this is the port being used to send the email(this specific one is only valid for gmail accounts)#
            smtp.ehlo()
            smtp.starttls()
            smtp.login('neavaccineappointment@gmail.com', 'DAvoGUNEA21/22')#this is where the email is being sent from; program has to login to account to send email#
            smtp.send_message(email)#sends email#
            print("Email sent!")#displays confirmation message when email sent#


        main_menu()#takes user back to main menu#
    def manage_appointment():#this is where the appointment detais can be edited or cancelled#
      try:
        full_name = input("Please enter your full name:")
      except:
        full_name = input("Invalid entry, please try again:")#ask the user to enter name again if entry invalid#
      try:
        enter_ref = int(input("Please enter your reference number: "))#asks user for reference number#
      except:
        enter_ref = int(input("Invalid entry, please try again:"))
        
      count = 1
      cursor.execute("SELECT reference FROM Appointments JOIN Clients WHERE Clients.id = ? ",[Appointment.ID])#retrieves reference number to check against user's entry#
      ref_num_format = repr(cursor.fetchone())#turns result of query from list to string#
      remove = ["'", "'", "[", ",", "]", "(", ")"]
      for ch in remove:
        ref_num_format = ref_num_format.replace(ch, "")#removes any unwanted characters from the string#
      ref_num_check = int(ref_num_format)
      
      if enter_ref != ref_num_check:
        while enter_ref != ref_num_check or count < 4:#while the user has not made 3 incorrect attempts#
          enter_ref = int(input("Incorrect input, try again:"))#asks user to enter again'#
          count += 1
        print("You have entered incorrectly too many times. You will now be returned back to the main menu.")
        main_menu()#after 3 incorrect attempts, user is returned back to main menu#
      else:
        print()
        print("Name: " + Appointment.name)#the user's appointment details are displayed#
        print("Postcode: " + Appointment.postcode)
        print("GP Postcode: " + Appointment.gp_pc)
        print("Appointment Date: " + str(Appointment.day))
        print("Appointment Time: " + str(Appointment.time) + ":00")
        print("Reference Number: " + str(Appointment.get_ref_num()))
        print()
        print("1. Edit information")
        print("2. Cancel appointment")
        print()
        choice = int(input("Please choose an option: "))#asks user to choose between editing info or cancelling appointment#
        if choice == 1:
          print()
          print("1. Name")
          print("2. Postcode")
          print("3. GP Postcode")
          print("4. Appointment Date")
          print("5. Appointment Time")
          print()
          edit = input("What would you like to change? ").lower()#asks user what they want to change if they choose to edit#
  
          if edit == "postcode":
            new_postcode = input("Please enter your new postcode:").upper()#asks user to enter new postcode#
            Appointment.postcode = new_postcode#assigns new postcode to postode attribute#
            print("You will now be returned back to the main menu.")              
            main_menu()#returns user to main menu#
          elif edit == "name":
            new_name = input("Please enter your new full name: ")#asks user to enter new name#
            Appointment.name = new_name#assigns new name to name attribute#
            print("You will now be returned back to the main menu.")            
            main_menu()#returns user back to main menu#
          elif edit == "gp postcode":#allows user to pick a new location#
            postcode = input("Enter your postcode: ").upper()
            Appointment.postcode = postcode
            f2 = str(postcode[0])
            ws = wb.active
            ans = {}
            count = 1
            for row in range(1,25):
                for col in range(3,4):
                    char = get_column_letter(col)
                    pc = ws[char + str(row)].value
                    match = re.search(r"\b"+f2,pc)
                    if match is not None:
                        ans[count] = pc
                        count = count + 1
            print(*[str(n)+ ":" + str(p) for n,p in ans.items()])
            
            
            try:
              pick_pc = int(input("Please pick a new location: "))
            except:
              pick_pc = int(input("Invalid entry, please try again:"))
            newuser_gp = ans[pick_pc]
            Appointment.gp_pc = newuser_gp
            print("You will now be returned back to the main menu.")
            main_menu()#returns user back to main menu#
          elif edit == "time":
            ClientTimetable = Timetable()
            NewTime = ClientTimetable.Time()#invokes time method from timetable class#
            Appointment.time = NewTime
            print("You will now be returned back to the main menu.")
            main_menu()#returns user back to main menu#
          elif edit == "date":
            ClientTimetable = Timetable()
            NewDate = ClientTimetable.Date()#invokes date method from timetable class#
            Appointment.day = NewDate
            print("You will now be returned back to the main menu.")            
            main_menu()#returns user back to main menu#         
  
        else:
          confirm = input("Are you sure you would like to cancel your appointment?(Y/N)").upper()#allows user to confirm option#
          if confirm == "Y":#if they confirm yes:#
            Appointment.name = "EMPTY"
            Appointment.postcode = "EMPTY"
            Appointment.gp_pc = "EMPTY"
            Appointment.time = "EMPTY"
            Appointment.day = "EMPTY"#then all the attributes are cleared#
            cursor.execute("DELETE FROM Appointments WHERE Appointments.id = ?", [Appointment.id])#all records which are related to the id given to the client are deleted from the database#
            print("You will now be returned back to the main menu.")
            main_menu()#user is returned back to main menu#
          else:
            print("You will now be returned to the main menu.")
            main_menu()#if they confirm no then user is returned back to main menu#
        while choice != "1" or choice != "2":
          choice = int(input("Invalid entry. Please enter again:"))#if user enters invalid option, then user has to enter again#
          

class Timetable:#relates to everything about the GP's timetable#
    def __init__(self):# all the attributes initialised to class#
        self.times = []
        self.times_available = self.times
        self.day = "EMPTY"
        self.time = "EMPTY"
        self.weekdays = []
        self.weekdays_editable = self.weekdays

    def Date(self):#method which returns client's appointment date#
        Today = datetime.date.today()
        Today_Format = Today.strftime("%x")#assigns todays date to variable#
        for days in range(1,8):#loops through rest of the week#
          NextDay = Today + datetime.timedelta(days = 1)
          NextDay_Format = NextDay.strftime("%x")
          self.weekdays.append(NextDay_Format)#adds next 6 days to weekdays attribute#
          Today = NextDay
        print()
        print(*self.weekdays_editable, sep = ", ")#displays dates to user#
        try:
          select_day = int(input("Enter the day you want your appointment to be booked for(day 0-6): "))#asks user to choose a date#
        except:
          select_day = int(input("Invalid entry, please try again(day 0-6): "))#asks user to enter again if an invalid entry is made#          

        self.day = self.weekdays_editable[select_day]#assigns user's choice to day attribute#
        if len(self.times_available) == 0:
          self.weekdays_editable.remove(self.day)#removes day from available days attribute#
        return self.day#returns day attribute to be used outside of method#

    def Time(self):#private method which returns client's appointment time#
      for times in range(9,18):
        self.times_available.append(times)#creates list with times between 9:00 to 17:00#
      print()
      print(*self.times_available, sep = ":00, ")#displays times to user#
      try:
        select_time = int(input("Please choose the time you would like your appointment to be at(0-6): "))#asks user to pick appointment time#
      except:
        select_time = int(input("Invalid entry, please try again(0-6): "))#asks user to choose again if there's an invalid entry#        
      self.time = self.times_available[select_time]#assigns user's to time attribute#
      self.times_available.remove(self.time)#removes time from available times attribute#
      return self.time#returns time attribute to be used outside of method#


def main_menu():#this is the main meun which will first be displayed to the user#
    print("************MAIN MENU***********")
    print()
    print("1. Book appointment")
    print("2. Manage appointment")
    print("3. Quit")
    print()
    try:
      choice = int(input("Please choose one of the options: "))#asks user to pick between booking or managing appointment#
      while choice < 1 or choice > 3:#asks user to enter again until valid option is entered#
        choice = int(input("Invalid entry. Please enter again: "))
    except:
      choice = int(input("Invalid entry. Please enter again: "))
      while choice < 1 or choice > 3:#asks user to enter again until valid option is entered#
        choice = int(input("Invalid entry. Please enter again: "))
    if choice == 1:
        new_app = Appointment.book_appointment()#instantiates book appointment method from Appointment class#
        
    elif choice == 2:
     manage = Appointment.manage_appointment()#instantiates manage appointment method from Appointment class#

    else:
      sys.exit("Thanks for using this service!")#quits the program in a user-friendly way#

main_menu()#calls main menu function#



#24/04/2022#
#Get project to remove time each time a user books an appointment#